import csv
import functools
import time
import typing
from typing import List, Tuple, Dict
from pathlib import Path

import PySide2
import numpy as np
from PySide2.QtCore import Qt, QCoreApplication, Signal, QModelIndex
from PySide2.QtWidgets import QWidget, QTableWidgetItem, QTableView, QHBoxLayout, QProgressDialog, QDialog, QMessageBox, \
    QMenu, QAction, QTextEdit, QVBoxLayout, QTabWidget, QTableWidget, QSizePolicy, QAbstractScrollArea, QLabel
import openpyxl

from arthropod_describer.common.blocking_operation import ProgressReport
from arthropod_describer.common.label_image import RegionProperty, PropertyType
from arthropod_describer.common.plugin import PropertyComputation, local_property_key, global_computation_key
from arthropod_describer.common.regions_cache import RegionsCache
from arthropod_describer.common.state import State
from arthropod_describer.common.units import convert_value, CompoundUnit, Unit, Value
from arthropod_describer.label_editor.computation_widget import ComputationWidget
from arthropod_describer.measurements_viewer.measurement_assign_dialog import MeasurementAssignDialog
from arthropod_describer.measurements_viewer.measurements_model import MeasurementsTableModel
from arthropod_describer.measurements_viewer.ui_measurements_viewer import Ui_MeasurementsViewer


class MeasurementsViewer(QWidget):
    open_project_folder = Signal()
    unsaved_changes = Signal()

    def __init__(self, state: State, parent: typing.Optional[PySide2.QtWidgets.QWidget] = None,
                 f: PySide2.QtCore.Qt.WindowFlags=Qt.WindowFlags()):
        super().__init__(parent, f)
        self.all_props: typing.List[typing.Tuple[str, str]] = []
        self.ui = Ui_MeasurementsViewer()
        self.ui.setupUi(self)

        self.state = state

        self.ui.btnComputeMeasurements.clicked.connect(self.show_new_measurements_dialog)
        self.ui.btnExport.clicked.connect(self.export_xlsx_results)
        self.ui.btnRecompute.clicked.connect(self.recompute)

        # TODO in future, a flexible export system would be nice
        self._export_menu = QMenu()
        self.csv_action = QAction("CSV")
        self.csv_action.triggered.connect(self.export_csv_results)
        self._export_menu.addAction(self.csv_action)

        self.ui.btnExport.setMenu(self._export_menu)

        self.computation_widget = ComputationWidget(self.state, None)
        self.computation_widget.set_settings_group_shown(False)
        self.computation_widget.set_restrict_group_shown(False)
        self.computation_widget.setVisible(False)
        self.measurement_assign_dialog = MeasurementAssignDialog(self.state,
                                                                 self.computation_widget.computations_model,
                                                                 parent=self)
        self.measurement_assign_dialog.setWindowModality(Qt.ApplicationModal)
        self.measurement_assign_dialog.ui.mainLayout.addWidget(self.computation_widget)
        self.setEnabled(False)

        self.prop_table_models: typing.Dict[str, MeasurementsTableModel] = {}

        #self.ui.results_widget.setLayout(QHBoxLayout())
        self.tables: typing.List[QTableView] = []

        self.model = MeasurementsTableModel(self.state)
        self.ui.chkColorVisually.toggled.connect(self.model.display_intensity_in_color)
        self.ui.tableView.doubleClicked.connect(self._handle_index_double_clicked)
        self.nd_browser = None

    def register_computation(self, comp: PropertyComputation):
        pass

    def show_new_measurements_dialog(self):
        # print(self.state.storage.used_regions('Labels'))
        to_compute: typing.Dict[str, typing.Set[int]] = self.measurement_assign_dialog.show_dialog()
        if len(to_compute) == 0:
            return

        all_labels: typing.Set[int] = set(functools.reduce(set().union, to_compute.values()))

        progress_dialog = ProgressReport(self.state.storage.image_count, 'Computing properties', self)

        for i in range(self.state.storage.image_count):
            photo = self.state.storage.get_photo_by_idx(i)
            regions_cache = RegionsCache(all_labels, photo, self.state.storage.default_label_image)
            for prop_path, labels in to_compute.items():
                region_props: List[RegionProperty] = []
                # for prop_path in prop_paths:
                comp_key, prop_key = global_computation_key(prop_path), local_property_key(prop_path)
                computation = self.computation_widget.computations_model.computations_dict[prop_path]
                # props_ = computation(photo, {label: {prop_key}}, regions_cache)
                props_ = computation(photo, labels, regions_cache)
                region_props.extend(props_)
                for prop in region_props:
                    photo[self.state.storage.default_label_image].set_region_prop(prop.label, prop)
            progress_dialog.increment()
        self.update_measurements_view()
        return

        photo_wise_assignments = [(i, to_compute) for i in range(self.state.storage.image_count)]
        computations: typing.Dict[str, typing.Dict[str, typing.Set[int]]] = {}
        computed_prop_names: typing.Set[str] = set()
        for label, prop_paths in to_compute.items():
            if len(prop_paths) == 0:
                continue
            for prop_path in prop_paths:
                prop_key = local_property_key(prop_path)
                comp_key = global_computation_key(prop_path)
                prop_labels = computations.setdefault(comp_key, {})
                prop_labels.setdefault(prop_key, set()).add(label)
                computation = self.computation_widget.computations_model.computations_dict[comp_key]
                computed_prop_names.add(f'"{computation.computes[prop_key].name}"')
        computed_prop_names: str = ', '.join(sorted(computed_prop_names))
        computed_prop_names: str = computed_prop_names.strip()
        if computed_prop_names[-1] == ',':
            computed_prop_names: str = computed_prop_names[:-1]
        progr_dialog = QProgressDialog(minimum=0, maximum=self.state.storage.image_count, parent=self)
        progr_dialog.setWindowModality(Qt.WindowModal)
        progr_dialog.setCancelButton(None)
        progr_dialog.setMinimumDuration(0)
        progr_dialog.setValue(0)
        progr_dialog.setLabelText('Preparing to compute properties...')
        progr_dialog.setWindowTitle('Computing properties')
        time.sleep(0.01)
        progr_dialog.setValue(0)
        QCoreApplication.processEvents()
        for i in range(self.state.storage.image_count):
        #for i, assignment in photo_wise_assignments:
            photo = self.state.storage.get_photo_by_idx(i)
            for computation_key, prop_labels in computations.items():
                computation: PropertyComputation = self.computation_widget.computations_model.computations_dict[computation_key]
                progr_dialog.setLabelText(f'Computing {computed_prop_names} for {photo.image_name}')
                progr_dialog.setValue(i + 1)
                reg_props = computation(photo, prop_labels)
                for prop in reg_props:
                    prop.info.key = f'{computation_key}.{prop.info.key}'
                    photo['Labels'].set_region_prop(prop.label, prop)
            progr_dialog.setValue(i + 1)
        progr_dialog.hide()
        self.update_measurements_view()

    def update_measurements_view(self):
        self.model.update_model()
        self.ui.tableView.setModel(None)
        self.ui.tableView.setModel(self.model)
        self.ui.tableView.selectionModel().selectionChanged.connect(self.enable_delete_recompute)
        self.ui.btnExport.setEnabled(True)
        self.ui.tableView.resizeColumnsToContents()
        #self.ui.tableView.reset()
        #for listwidget in self.tables:
        #    self.ui.results_widget.layout().removeWidget(listwidget)
        #    listwidget.deleteLater()
        #self.tables.clear()

        #all_props: typing.Set[typing.Tuple[str, str]] = set()
        #for i in range(self.state.storage.image_count):
        #    photo = self.state.storage.get_photo_by_idx(i, load_image=False)
        #    for prop in photo['Labels'].prop_list:
        #        all_props.add((prop.info.name, prop.info.key))

        #self.all_props: typing.List[typing.Tuple[str, str]] = list(all_props)
        #self.all_props.sort()

        #for column in range(len(self.all_props)):
        #    #header_item = QTableWidgetItem(text=self.all_props[column][0])
        #    #header_item.setToolTip(self.all_props[column][1])
        #    #self.ui.measurementsTable.setHorizontalHeaderItem(column, header_item)
        #    model = MeasurementsTableModel(self.state, None)
        #    model.display_property(self.all_props[column][1])
        #    self.prop_table_models[self.all_props[column][1]] = model
        #    table_view = QTableView(None)
        #    self.tables.append(table_view)
        #    table_view.setModel(model)
        #    self.ui.results_widget.layout().addWidget(table_view)
        #    #index = self.ui.measurementsTable.model().index(0, column)
        #    #self.ui.measurementsTable.setIndexWidget(index, table_view)

    def export_csv_results(self):
        nd_props, other_props = self._filter_by_NDArray(self.model.prop_tuple_list)

        file_grouped_props = self._group_measurements_by_sheet(other_props)

        file_names: typing.List[str] = []

        for group_name, prop_list in file_grouped_props.items():
            with open(self.state.storage.location / f'results_{group_name}.csv', 'w', newline='') as f:
                writer = csv.writer(f, dialect='excel')
                self._tabular_export_routine(prop_list, writer.writerow)
            file_names.append(f'results_{group_name}.csv')

        sheet_nd_props = self._group_measurements_by_sheet(nd_props)

        for group_name, prop_list in sheet_nd_props.items():
            with open(self.state.storage.location / f'results_{group_name}.csv', 'w', newline='') as f:
                writer = csv.writer(f, dialect='excel')
                self._ndarray_export_routine(prop_list, writer.writerow)
            file_names.append(f'results_{group_name}.csv')

        # filenames = '\n'.join(file_names)
        # filenames = filenames[:-2]
        self.show_export_success_message(self.state.storage.location, file_names)

    def export_xlsx_results(self):
        wb = openpyxl.Workbook()

        # self._tabular_export_routine(ws.append)
        nd_props, other_props = self._filter_by_NDArray(self.model.prop_tuple_list)

        sheet_grouped_props = self._group_measurements_by_sheet(other_props)

        for sheet_name, prop_list in sheet_grouped_props.items():
            ws = wb.create_sheet(sheet_name)
            self._tabular_export_routine(prop_list, ws.append)

        sheet_nd_props = self._group_measurements_by_sheet(nd_props)

        for sheet_name, prop_list in sheet_nd_props.items():
            ws = wb.create_sheet(sheet_name)
            self._ndarray_export_routine(prop_list, ws.append)

        if 'Sheet' in wb:
            wb.remove(wb['Sheet'])

        wb.save(str(self.state.storage.location / 'results.xlsx'))

        self.show_export_success_message(self.state.storage.location, ['results.xlsx'])

    def _ndarray_export_routine(self, prop_list: List[Tuple[int, str, str]], append_row):
        row = []
        prop_list = list(sorted(prop_list, key=lambda tup: tup[0]))
        for i in range(self.state.storage.image_count):
            photo = self.state.storage.get_photo_by_idx(i, load_image=False)
            lab_img = photo['Labels']
            row = [photo.image_name]
            for label, prop_key, prop_name in prop_list:
                if lab_img.get_region_props(label) is None:
                    continue
                reg_props = lab_img.get_region_props(label)

                if prop_key not in reg_props:
                    continue
                prop = reg_props[prop_key]
                region_name = lab_img.label_hierarchy.nodes[label].name

                for j in range(prop.num_vals):
                    if j == 0:
                        row.append(f'{region_name}:{prop.info.name} {prop.val_names[j]}')
                    else:
                        row = ['', f'{region_name}:{prop.info.name} {prop.val_names[j]}']
                    for col in prop.col_names:
                        row.append(col)
                    append_row(row)
                    matrix2d: np.ndarray = prop.value[0][j]
                    for r in range(matrix2d.shape[0]):
                        row = ['', prop.row_names[r]]
                        for c in range(matrix2d.shape[1]):
                            row.append(f'{matrix2d[r, c]:.3f}')
                        append_row(row)
                    row = ['']
                    append_row(row)
                append_row([''])

    def _group_measurements_by_sheet(self, prop_tup_list: List[Tuple[int, str, str]]) \
            -> typing.Dict[str, typing.List[typing.Tuple[int, str, str]]]:
        sheet_grouped_props: typing.Dict[str, typing.List[typing.Tuple[int, str, str]]] = {}

        for label, prop_key, prop_name in prop_tup_list:
            dot_split = prop_key.split('.')
            comp_key = '.'.join(dot_split[:-1])
            computation = self.computation_widget.computations_model.computations_dict[prop_key]
            prop_key_local = dot_split[-1]
            sheet_grouped_props.setdefault(computation.target_worksheet(prop_key_local), []).append((label,
                                                                                                     prop_key,
                                                                                                     prop_name))

        return sheet_grouped_props

    def _filter_by_NDArray(self, prop_tuple_list: List[Tuple[int, str, str]]) -> \
        Tuple[List[Tuple[int, str, str]], List[Tuple[int, str, str]]]:
        ndarray_props: List[Tuple[int, str, str]] = []
        other_props: List[Tuple[int, str, str]] = []

        for label, prop_key, prop_name in prop_tuple_list:
            dot_split = prop_key.split('.')
            comp_key = '.'.join(dot_split[:-1])
            # computation = self.computation_widget.computations_model.computations_dict[comp_key]
            computation = self.computation_widget.computations_model.computations_dict[prop_key]
            if computation.example(dot_split[-1]).prop_type == PropertyType.NDArray:
                ndarray_props.append((label, prop_key, prop_name))
            else:
                other_props.append((label, prop_key, prop_name))
        return ndarray_props, other_props

    def show_export_success_message(self, folder: Path, filenames: List[str]):
        filenames_in_rows = '\n'.join(filenames)
        if QMessageBox.information(self, 'Export finished',
                                   f'The results were saved in the folder {folder} as files:\n{filenames_in_rows}\nDo you want to open the folder?',
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No) == QMessageBox.Yes:
            self.open_project_folder.emit()

    def _tabular_export_routine(self, prop_tuple_list: typing.List[typing.Tuple[int, str, str]],
                                append_row: typing.Callable[[typing.List[typing.Any]], None]):
        example_props: typing.Dict[str, RegionProperty] = {}
        column_names = ['photo_name']
        for label, prop_key, prop_name in prop_tuple_list:
            dot_split = prop_key.split('.')
            comp_key = '.'.join(dot_split[:-1])
            computation = self.computation_widget.computations_model.computations_dict[prop_key]
            prop = computation.example(dot_split[-1])
            example_props[prop_key] = prop
            # if prop.prop_type == PropertyType.Vector or prop.num_vals > 1:
            if prop.num_vals > 1:
                if len(prop.val_names) != prop.num_vals:
                    col_names = [str(i) for i in range(prop.num_vals)]
                else:
                    col_names = prop.val_names
                for i in range(prop.num_vals):
                    #column_names.append(f'{prop.info.key}_{prop.val_names[i]}:{self.state.colormap.label_names[label]}')
                    # column_names.append(f'{prop.info.key}_{prop.val_names[i]}:{self.state.label_hierarchy.nodes[label].name}')
                    column_names.append(f'{self.state.label_hierarchy.nodes[label].name}:{prop.info.name}_{col_names[i]}')
            else:
                #column_names.append(f'{prop.info.key}:{self.state.colormap.label_names[label]}')
                # column_names.append(f'{prop.info.key}:{self.state.label_hierarchy.nodes[label].name}')
                column_names.append(f'{self.state.label_hierarchy.nodes[label].name}:{prop.info.name}')
        append_row(column_names)
        for i in range(self.state.storage.image_count):
            photo = self.state.storage.get_photo_by_idx(i, load_image=False)
            row = [photo.image_name]
            label_img = photo['Labels']
            for label, prop_key, prop_name in prop_tuple_list:
                if label_img.get_region_props(
                        label) is None:  # So this LabelImg does not have props for `label`, insert a sequence of 'missing value' symbols (-1)
                    ex_prop = example_props[prop_key]
                    for _ in range(ex_prop.num_vals):
                        row.append('N/A')
                else:
                    reg_props = label_img.get_region_props(label)
                    if prop_key not in reg_props:  # This label region does not have property with `prop_key`, insert a sequence of `missing value` symbols (-1)
                        ex_prop = example_props[prop_key]
                        for _ in range(ex_prop.num_vals):
                            row.append('N/A')
                    else:  # Finally, insert actual values for present property
                        reg_prop: RegionProperty = reg_props[prop_key]
                        if reg_prop.num_vals > 1:
                            unit_: typing.Union[Unit, CompoundUnit] = reg_prop.value[1]
                            targ_unit = self.state.units.get_default_unit(unit_)
                            # mult = (10 ** (int(reg_prop.value[1].prefix) - int(targ_unit.prefix))) ** reg_prop.value[1].dim
                            val_ = Value(reg_prop.value[0][0], unit_)
                            for val in reg_prop.value[0]:
                                val_.value = val
                                n_val = convert_value(val_, targ_unit)
                                row.append(n_val.value)
                                # break  # TODO how to handle exporting vector of values to CSV?
                        else:
                            targ_unit = self.state.units.get_default_unit(reg_prop.value.unit)
                            conv_val = convert_value(reg_prop.value, targ_unit)
                            row.append(conv_val.value)
            append_row(row)

    def enable_delete_recompute(self):
        enable = len(self.ui.tableView.selectedIndexes()) > 0
        self.ui.btnDelete.setEnabled(enable)
        self.ui.btnRecompute.setEnabled(enable)

    def recompute(self):
        idxs = self.ui.tableView.selectedIndexes()
        row_wise: typing.List[QModelIndex] = list(sorted(idxs, key=lambda idx: idx.row()))
        assignments: typing.List[typing.Tuple[int, typing.Dict[str, typing.Set[int]]]] = []
        assignment: typing.Dict[str, typing.Set[int]] = {}
        photo_idx = row_wise[0].row()
        for idx in row_wise:
            if idx.row() != photo_idx:
                assignments.append((photo_idx, assignment))
                photo_idx = idx.row()
                assignment = {}
            label = self.model.data(idx, Qt.UserRole)
            prop_key = self.model.data(idx, Qt.UserRole + 1)
            assignment.setdefault(prop_key, set()).add(label)
        assignments.append((row_wise[-1].row(), assignment))
        self.compute_measurements(assignments)

    def compute_measurements(self, photo_assignments: typing.List[typing.Tuple[int, typing.Dict[str, typing.Set[int]]]]):
        progr_dialog = QProgressDialog(minimum=0, maximum=len(photo_assignments), parent=self)
        progr_dialog.setWindowModality(Qt.WindowModal)
        progr_dialog.setCancelButton(None)
        progr_dialog.setMinimumDuration(0)
        progr_dialog.setValue(0)
        progr_dialog.setLabelText('Preparing to compute properties...')
        progr_dialog.setWindowTitle('Computing properties')
        time.sleep(0.01)
        progr_dialog.setValue(0)
        QCoreApplication.processEvents()
        for progress_value, (i, to_compute) in enumerate(photo_assignments):
            photo = self.state.storage.get_photo_by_idx(i)
            computations = {}
            # for prop_key, labels in to_compute.items():
            #     dot_splits = prop_key.split('.')
            #     prop_name = dot_splits.pop()
            #     comp_key = '.'.join(dot_splits)
            #     prop_labels = computations.setdefault(prop_key, labels)
            #     # prop_labels[prop_name] = labels
            all_labels = set(functools.reduce(set.union, to_compute.values()))
            regs_cache = RegionsCache(all_labels, photo, 'Labels')
            for prop_key, prop_labels in to_compute.items():
                computation: PropertyComputation = self.computation_widget.computations_model.computations_dict[prop_key]
                progr_dialog.setLabelText(f'Computing {computation.info.name} for {photo.image_name}')
                progr_dialog.setValue(progress_value + 1)
                reg_props = computation(photo, list(prop_labels), regs_cache)
                for prop in reg_props:
                    # prop.info.key = f'{computation_key}.{prop.info.key}'
                    prop.info.key = computation.info.key
                    photo['Labels'].set_region_prop(prop.label, prop)
            progr_dialog.setValue(progress_value + 1)
        progr_dialog.hide()
        self.update_measurements_view()
        self.unsaved_changes.emit()

    def _handle_index_double_clicked(self, index: QModelIndex):
        prop: RegionProperty = self.model.data(index, Qt.UserRole + 4)
        if prop.prop_type != PropertyType.NDArray:
            return
        if self.nd_browser is not None:
            size = self.nd_browser.size()
        else:
            size = None
        self.nd_browser = QTabWidget()

        table_widget = QWidget()
        table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        tables_layout = QVBoxLayout()
        table_widget.setLayout(tables_layout)

        text = ''

        np.set_printoptions(precision=3)
        for i in range(prop.num_vals):
            nd: np.ndarray = prop.value[0][i]
            if nd.ndim > 2:
                raise ValueError("Not implemented for ndim > 2")
            elif nd.ndim == 1:
                nd = np.array(nd)
            table = QTableWidget()
            table.setRowCount(nd.shape[0])
            table.setColumnCount(nd.shape[1])
            table.setHorizontalHeaderLabels(prop.col_names)
            table.setVerticalHeaderLabels(prop.row_names)

            for r in range(nd.shape[0]):
                for c in range(nd.shape[1]):
                    titem = QTableWidgetItem(f'{nd[r, c]:.3f}')
                    table.setItem(r, c, titem)
            tables_layout.addWidget(QLabel(prop.val_names[i]))
            tables_layout.addWidget(table)
            table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
            # table.adjustSize()

            text += f'\n{prop.val_names[i]}\n'
            text += str(prop.value[0][i])
            text += '\n'

        self.nd_browser.addTab(table_widget, 'Table view')

        self.nd_browser.setWindowTitle(f'{self.state.label_hierarchy.nodes[prop.label].name} - {prop.info.name}')
        lay = QVBoxLayout()
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.zoomIn(3)
        text_edit.setText(text)
        self.nd_browser.addTab(text_edit, 'Raw view')
        self.nd_browser.setWindowModality(Qt.ApplicationModal)
        self.nd_browser.show()
        if size is not None:
            self.nd_browser.setFixedSize(size)
        table_widget.adjustSize()


