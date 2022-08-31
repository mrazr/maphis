import csv
import itertools
import shutil
from pathlib import Path
import typing
from typing import List, Optional, Set

import numpy as np
from skimage import io
import openpyxl
# import matplotlib.pyplot as plt

from arthropod_describer.common.common import Info
from arthropod_describer.common.label_image import LabelImg, RegionProperty
from arthropod_describer.common.photo import Photo

from arthropod_describer.common.plugin import GeneralAction
from arthropod_describer.common.state import State
from arthropod_describer.common.storage import Storage
from arthropod_describer.common.user_params import UserParam
from arthropod_describer.plugins.profile_register.general.profiles import get_median_profile, merge_profiles


class ProfileFusion(GeneralAction):
    """
    NAME: Profile fusion
    DESCRIPTION: Fuse body profiles based on their tags. A median profile is created for each group of images that
    matches the given tags. (how to specify and use the tags must be found)

    USER_PARAMS:
        PARAM_NAME: Iteration tag
        PARAM_KEY: iteration_tags
        PARAM_SOURCE: Storage
        PARAM_SOURCE_FIELD: tag_prefixes
        PARAM_VALUE_CARDINALITY: SingleValue

        PARAM_NAME: Model group tags
        PARAM_KEY: model_group_tags
        PARAM_SOURCE: Storage
        PARAM_SOURCE_FIELD: tag_prefixes
        PARAM_VALUE_CARDINALITY: MultiValue

        PARAM_NAME: Mimic group tags
        PARAM_KEY: mimic_group_tags
        PARAM_SOURCE: Storage
        PARAM_SOURCE_FIELD: tag_prefixes
        PARAM_VALUE_CARDINALITY: MultiValue

        PARAM_NAME: Delete existing xlsx files
        PARAM_KEY: delete_existing_xlsx
        PARAM_TYPE: BOOL
        DEFAULT_VALUE: FALSE
    """
    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)

    def __call__(self, state: State):
        # state.storage
        # state.storage.image_names
        # image = state.storage.get_photo_by_name()
        # extract model and mimic image names from storage.image_names
        # get model and mimic images by calling storage.get_photo_by_name(<image_name>)
        # for each `image` the contour vector is stored in `image['Labels'].region_props[16842752]['contour'].value`,
        # that is of type `RegionProperty`, and the vector is stored in `RegionProperty.value`, and in the case of
        # the contour vector, it is a List[float]
        # print(f'model prefix is {self._user_params["iteration_tags"].value}, mimic prefix is {self._user_params["group_tags"].value}')

        output_folder = Path(state.storage.location / 'registered_profiles')
        if not output_folder.exists():
            output_folder.mkdir(exist_ok=True)

        print(f'iteration tags are {self._user_params["iteration_tags"].param_instances[0].value}')

        print(f'model group tags are: {self._user_params["model_group_tags"].value}')
        print(f'mimic group tags are: {self._user_params["mimic_group_tags"].value}')
        # for group_tags in self._user_params['group_tags'].param_instances.values():
        #     print(group_tags.value)

        prefix_tags_map: typing.Dict[str, typing.List[str]] = dict()
        all_storage_tags = state.storage.used_tags

        for tag_prefix in self._user_params['iteration_tags'].value:
            for tag in all_storage_tags:
                if tag.startswith(tag_prefix):
                    prefix_tags_map.setdefault(tag_prefix, list()).append(tag)
        print(f'prefix_tags_map = {prefix_tags_map}')
        # print(f'iteration tag sets are {list(itertools.product(*prefix_tags_map.values()))}')

        iteration_tags = [tag for tag in all_storage_tags if tag.startswith(self._user_params["iteration_tags"].value)]
        print(f'iteration tags are {iteration_tags}')

        model_tags = self._user_params["model_group_tags"].value
        model_column_name = "_".join(sorted(model_tags))

        mimic_tags = self._user_params["mimic_group_tags"].value
        mimic_column_name = "_".join(sorted(mimic_tags))
        registered_column_name = model_column_name + ";" + mimic_column_name

        if self._user_params['delete_existing_xlsx'].value:
            for iteration_tag in iteration_tags:
                shutil.rmtree(output_folder / iteration_tag, ignore_errors=True)

        for iteration_tag in iteration_tags:
            sample_folder = output_folder / iteration_tag
            if not sample_folder.exists():
                sample_folder.mkdir(exist_ok=True)
            if (worksheet_path := sample_folder / 'profiles.xlsx').exists():
                wb = openpyxl.load_workbook(worksheet_path, read_only=False)
            else:
                wb = openpyxl.Workbook()
                ws_med_prof = wb.active
                ws_med_prof.title = 'Median profiles'
                ws_med_prof.append(['ProfileID'] + [f'G_BP_{i}' for i in range(40)])

                ws_aligned = wb.create_sheet('Aligned profiles')
                ws_aligned.append(['ProfileID', 'AlignedTo'] + [f'G_BP_{i}' for i in range(40)])

            ws_med_prof = wb['Median profiles']

            model_median = self._get_median_profile_for(iteration_tag, model_tags, state)
            mimic_median = self._get_median_profile_for(iteration_tag, mimic_tags, state)

            ws_med_prof.append([model_column_name] + [str(float(val)) for val in model_median])
            ws_med_prof.append([mimic_column_name] + [str(float(val)) for val in mimic_median])

            # if model_median is not None:
            #     np.savetxt(str(sample_folder / f'{model_column_name}.txt'), model_median)
            # if mimic_median is not None:
            #     np.savetxt(str(sample_folder / f'{mimic_column_name}.txt'), mimic_median)

            if model_median is None or mimic_median is None:
                continue  # TODO log a message

            # registered = merge_profiles(model_median, mimic_median, show_matches=True)
            # np.savetxt(str(sample_folder / f'{registered_column_name}.txt'), registered)

            # model_mimic_1 = merge_profiles(model_median, mimic_median, x_weight=1.0, y_weight=1.0)
            model_mimic_05 = merge_profiles(model_median, mimic_median)
            # model_mimic_0 = merge_profiles(model_median, mimic_median, x_weight=0.0, y_weight=0.0)

            # mimic_aligned_to_model = merge_profiles(mimic_median, model_mimic_1)
            # mimic_aligned_to_mimic = merge_profiles(mimic_median, model_mimic_0)
            mimic_aligned_to_both = merge_profiles(mimic_median, model_mimic_05)

            # model_aligned_to_model = merge_profiles(model_median, model_mimic_1)
            # model_aligned_to_mimic = merge_profiles(model_median, model_mimic_0)
            model_aligned_to_both = merge_profiles(model_median, model_mimic_05)

            # plt.subplots(nrows=1, ncols=2)
            # plt.subplot(121)
            # plt.plot(range(40), model_aligned_to_both)
            #
            # plt.subplot(122)
            # plt.plot(range(40), mimic_aligned_to_both)
            #
            # plt.savefig('D:/plot.png')
            # plt.show()

            ws_aligned = wb['Aligned profiles']

            # ws_aligned.append([mimic_column_name, model_column_name] + [str(float(val)) for val in mimic_aligned_to_model])
            # ws_aligned.append([mimic_column_name, mimic_column_name] + [str(float(val)) for val in mimic_aligned_to_mimic])

            ws_aligned.append([model_column_name, f'{model_column_name}_{mimic_column_name}_0.5_0.5'] + [str(float(val)) for val in model_aligned_to_both])

            ws_aligned.append([mimic_column_name, f'{model_column_name}_{mimic_column_name}_0.5_0.5'] + [str(float(val)) for val in mimic_aligned_to_both])

            # ws_aligned.append([model_column_name, model_column_name] + [str(float(val)) for val in model_aligned_to_model])
            # ws_aligned.append([model_column_name, mimic_column_name] + [str(float(val)) for val in model_aligned_to_mimic])

            wb.save(worksheet_path)
            wb.close()

            # with open(sample_folder / f'{registered_column_name};profiles.csv', newline='\n', mode='w') as csvfile:
            #     writer = csv.writer(csvfile, dialect='excel')
            #     writer.writerow([model_column_name, mimic_column_name, registered_column_name])
            #
            #     for mod_val, mim_val, reg_val in zip(model_median, mimic_median, registered):
            #         writer.writerow([mod_val, mim_val, reg_val])

    def _get_median_profile_for(self, iteration_tag: str, group_tags: typing.Set[str], state: State) -> typing.Optional[np.ndarray]:
        tags_set = set(group_tags).union({iteration_tag})

        photos = state.storage.photos_satisfying_tags(tags_set)

        if len(photos) == 0:
            return None
        profiles: np.ndarray = np.zeros((len(photos), 40), np.float32)
        print(f'number of photos satisfying the tag set {tags_set} = {len(photos)}')
        unit = None
        for i, photo in enumerate(photos):
            # io.imsave(str(save_dest / photo.image_name), photo.image)
            profiles[i] = np.array(photo['Labels'].region_props[16842752]["arthropod_describer.plugins.profile_register.properties.contour"].value[0])
            # unit = photo['Labels'].region_props[16842752]["arthropod_describer.plugins.profile_register.properties.contour"].value[1]
        median = get_median_profile(profiles, show_fig=False)
        # prop = RegionProperty()
        # prop.label = 16842752
        # prop.num_vals = profiles.shape[1]
        # prop.value = (median, unit)
        # state.storage.properties[self.info.key]['_'.join(tags_set)] = prop
        # np.save('D:/profiles_registering_test/median.npy', median)
        # np.savetxt(f'median{"_".join(tags_set)}.txt', median)
        return median

    @property
    def user_params(self) -> List[UserParam]:
        return super().user_params
